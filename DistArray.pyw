import openpyxl
from openpyxl import load_workbook
import clipboard
# store the clipboard to a variable
clip = clipboard.paste()
# load workbook and set list_1 worksheet & column A as variables
wb = load_workbook(filename = "W:\CC_FS\CCC\Data In\DIS Team\BrandMapping_RPG Assignment\RPG Assignment2.1.xlsm")
sheet = wb["LIST_1"]
columnA = sheet["A"]
# loop trough column A and search for the desired code. Return cells on the same row.
for cell in columnA:
    if cell.value == int(clip):
        cl = cell.row
        c1 = cell.column +1
        c2 = cell.column +2
        c3 = cell.column +3
        c4 = cell.column +4
# put the row results into the clipboard and send it to the .ahk script to display
output = [sheet.cell(row=cl, column=cell.column).value , sheet.cell(row=cl, column=c1).value, sheet.cell(row=cl, column=c2).value, sheet.cell(row=cl, column=c3).value, sheet.cell(row=cl, column=c4).value]
clipboard.copy(str(output))
input()